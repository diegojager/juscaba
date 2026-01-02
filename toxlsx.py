import openpyxl 
from openpyxl.styles import Font
import datetime, decimal
 
def armarTitulos(ws:openpyxl.Workbook.active , campos:list , primeraColumna:int = 0):
    #ws.append([campo for campo in campos])
    ft = Font(bold=True)
    for i in range(1,len(campos)+1):
        ws.cell(row=1, column=primeraColumna + i).value=campos[i-1]
        ws.cell(row=1, column=primeraColumna + i).font=ft
 
def llenarCelda(ws, i, j, valor):
    ws.cell(row = i, column = j, value = valor)
    if type(valor) == decimal.Decimal:
        ws.cell(i,j).number_format="0.00"
    if type(valor) == int:
        ws.cell(i,j).number_format="0"
    elif type(valor) == datetime.date:
        ws.cell(i,j).number_format='dd/mm/yyyy'


def pipeCSVtoXlsx(listofstrings, campos, nombrearchivo, primeraColumna:int = 0, archivoExistente:bool = False):
    if archivoExistente:
        wb = openpyxl.load_workbook(filename=nombrearchivo)
    else:
        wb = openpyxl.Workbook()
    ws = wb.active
    armarTitulos(ws, campos, primeraColumna)
    for fila in listofstrings:
        lista = fila.split('|')
        j = 1
        for valor in lista[1:]:
            llenarCelda(ws,int(lista[0]),j+primeraColumna,valor)
            j += 1
    wb.close()
    wb.save(nombrearchivo)

def querySetToXlsx(qs, campos, nombrearchivo):
    wb = openpyxl.Workbook()
    ws = wb.active
    armarTitulos(ws, campos)

    for fila, registro in enumerate(qs, start=1):
        for columna, campo in enumerate(campos,start=1):
            #valor = registro[campo[0]]
            llenarCelda(ws, fila+1, columna, registro[campo[0]])
    wb.close()
    return (wb, nombrearchivo)

def completarDatosXlsx():
    return True


def encontrarUltimaColumna(wbx:openpyxl.Workbook = None , columnaMaxima:int =None, nombreArchivo:str = None):
    wb =  openpyxl.load_workbook(nombreArchivo) if nombreArchivo is not None else wbx
    columnaMaxima = hoja.max_column+1 if columnaMaxima is None else columnaMaxima
    nombreUltimaColumna = None
    indiceUltimaColumna = 0
    hoja = wb.active 
    
    for i in range(columnaMaxima, 0, -1):
        v  = hoja.cell(1, i ).value
        if ( v is not None):
            indiceUltimaColumna = i
            nombreUltimaColumna = v            
            break
    if nombreArchivo is not None:
        wb.close()
        
    return (indiceUltimaColumna, nombreUltimaColumna)

def testAppend(wbName, columnas):
    wb = openpyxl.load_workbook(filename=wbName)
    nroUltimaColumna, nombreUltimaColumna  = encontrarUltimaColumna(wb, 255)
    dictCnas = {}
    for i in range(len(columnas)):
        dictCnas [nroUltimaColumna + i + 1] = columnas[i]
        wb.active.cell(1, nroUltimaColumna+i+1 ).value = columnas[i] 
    
    wb.close()
    wb.save(wbName)
    
    return True

if __name__ == "__main__":
    import sys
    import openpyxl
    columnas = ['uno','dos','tres']
    testAppend(sys.argv[1], columnas)