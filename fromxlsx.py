import openpyxl 
import sys
import unicodedata
import re
import datetime, decimal


def leerTitulos(hoja: openpyxl.worksheet):
    titulos = []
    for i in range(1,hoja.max_column+1):
        if (isinstance(hoja.cell(1,i).value,str)):
            titulos.append(unicodedata.normalize('NFKD', hoja.cell(1,i).value.lower()).encode("ascii","ignore").decode("ascii"))
    return titulos

def leerColumna(hoja: openpyxl.worksheet, columnaAdjudicacion:int, columnaJuzgado: int):
    valores = {}
    #x =  hoja.cell(2,columnaJuzgado).value
    for i in range(2, hoja.max_row+1):
        if columnaJuzgado:
            vj= hoja.cell(i,columnaJuzgado).value
            if vj is None or int(vj) == 0:
                va = hoja.cell(i, columnaAdjudicacion).value
                if va is not None:
                    valores[i]=va
    return valores
    
def leerAdjudicacionDesdeXlsx(nombre: str):
    wb = openpyxl.load_workbook(filename=nombre)
    columnas = leerTitulos(wb.active)
    try:
        columnaAdjudicacion = columnas.index("adjudicacion")+1
    except ValueError:
        print(f"Falta columna adjudicacion en {nombre}")
        return []
    try:
        columnaJuzgado = columnas.index("juzgado")+1
    except ValueError:
        columnaJuzgado = None
    ret = leerColumna(wb.active, columnaAdjudicacion, columnaJuzgado)
    wb.close()
    return ret



if __name__ == "__main__":
    leerAdjudicacionDesdeXlsx(sys.argv[1])
    